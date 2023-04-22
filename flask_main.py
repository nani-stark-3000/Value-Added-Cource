from flask import Flask,render_template,request,make_response
import xlsxwriter
import openpyxl
import pandas as pd
import secret as s
import boto3
import cv2
import os
import io

app = Flask(__name__,template_folder='template')

s3 = boto3.resource(
    service_name='s3',
    region_name="ap-south-1",
    aws_access_key_id= s.access_key,
    aws_secret_access_key= s.secret_access_key
)

os.environ["AWS_DEFAULT_REGION"] = s.region_name
os.environ["AWS_ACCESS_KEY_ID"] = s.access_key
os.environ["AWS_SECRET_ACCESS_KEY"] = s.secret_access_key


def update(course_name):
    s3.Bucket('vaccourse').upload_file(Filename=course_name, Key=course_name)

def create(course_name,year,branch):
    print("Enter details to create sheet... ")
    filename = course_name+'-'+branch+'-'+year+'.xlsx'
    print(filename,type(filename))
    wb = xlsxwriter.Workbook(filename)
    sheet1 = wb.add_worksheet("Attendance")
    sheet2 = wb.add_worksheet("Feedback")
    sheet1.write(0,0,'Name')
    sheet1.write(0,1,'Reg Number')
    sheet1.write(0,2,'Day-1')
    sheet2.write(0,1,'Course Content')
    sheet2.write(0,2,'Coverage Of Syllabus')
    sheet2.write(0,3,'Very Helpful For Your Skill Development')
    sheet2.write(0,4,'Effectiveness Of The Course')
    sheet2.write(0,5,'Interatcion & Individual Attenction')
    sheet2.write(1,0,'Excelent')
    sheet2.write(2,0,'Very Good')
    sheet2.write(3,0,'Good')
    sheet2.write(4,0,'Satisfactory')
    sheet2.write(5,0,'To Be Improved')
    for i in range(1,6):
        for j in range(1,6):
            sheet2.write(i,j,0)
    wb.close()
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    update(filename)
    return filename,sheet

def register(course_name,name,reg_number):
    wb = openpyxl.load_workbook(course_name)
    ws = wb["Attendance"]
    info = (name,reg_number)
    ws.append(info)
    wb.save(course_name)
    update(course_name)
    return ws

def course_list():
    list = []
    for obj in s3.Bucket('vaccourse').objects.all():
        list.append(obj.key)
    return list

def select(s_course):
    s3.Bucket('vaccourse').download_file(Key=s_course, Filename=s_course)

def a_daylist(course_name):
    df = pd.read_excel(course_name)
    col = list(df.columns)
    alpha ='A'
    day = {'add':0}
    for i in col:
        day[i]=alpha
        alpha=chr(ord(alpha)+1)
    day_list = [i for i in day]
    day_list.append('add')
    return day,day_list[3:]

def s_daylist(course_name):
    df = pd.read_excel(course_name)
    col = list(df.columns)
    alpha ='A'
    day = {}
    for i in col:
        day[i]=alpha
        alpha=chr(ord(alpha)+1)
    day_list = [i for i in day]
    return day,day_list[2:]

def add_day(day,course_name):
    wb = openpyxl.load_workbook(course_name)
    ws = wb['Attendance']
    new_day = 'Day-'+str(len(day)-2)
    prev_day = 'Day-'+str(len(day)-3)
    day[new_day] = chr(ord(day[prev_day])+1)
    ws[day[new_day]+'1']=new_day
    print("added",new_day)
    return day[new_day],new_day

def a_post(choise,day,course_name,nlist,llist):
    today = day[choise]
    if today == 0 :
        today,choise=add_day(day,course_name)
    head = today+'1'
    wb = openpyxl.load_workbook(course_name)
    sheet = wb.active
    sheet[head]= choise

    n= ['21L31A54'+i for i in nlist]
    l= ['22L35A54'+i for i in llist]
    rows = sheet.max_row
    for i in range(2,rows+1):
        roll='B'+str(i)
        atnd=today+str(i)
        if (sheet[roll].value in n) or (sheet[roll].value in l):
            sheet[atnd]= 0
        else:
            sheet[atnd]= 1 
    wb.save(course_name)
    update(course_name)
    return sheet

def s_post(choise,day,course_name,roll_no,status):
    today = day[choise]
    wb = openpyxl.load_workbook(course_name)
    sheet = wb.active
    rows = sheet.max_row
    for i in range(2,rows+1):
        roll='B'+str(i)
        atnd=today+str(i)
        if sheet[roll].value==roll_no:
            sheet[atnd]= status
            print("Attendance updated ")
            break
    wb.save(course_name)
    update(course_name)
    return sheet

def a_status(course_name,roll_no):
    wb = openpyxl.load_workbook(course_name)
    sheet = wb.active
    rows = sheet.max_row
    col = sheet.max_column
    status=flag = 0
    for i in range(2,rows+1):
        roll='B'+str(i)
        if sheet[roll].value==roll_no:
            flag=1
            alpha = 'C'
            for j in range(3,col+1):
                if int(sheet[alpha+str(i)].value)==1:
                    status+=1
                alpha = chr(ord(alpha)+1)
    wb.save(course_name)
    update(course_name)
    if flag:
        percentage = (status/(col-2))*100
        comment='Your attendance is '+str(status)+' for '+str(col-2)+' days and your percentage is '+str(percentage)+'%'
        return comment,sheet
    else:
        comment= 'data not found with ..'+str(roll_no)
        return comment,sheet

@app.route('/')
def _home_():
    return render_template("home.html")

@app.route('/home')
def _home1_():
    return render_template("home.html")

@app.route('/create',methods=['POST','GET'])

def _create_():
    if request.method=='POST':
        course_name = request.form['course']
        branch = request.form['branch']
        year = request.form['year']
        file,sheet = create(course_name,year,branch)
        return render_template("create.html",file=file,sheet=sheet)
    return render_template("create.html")

@app.route('/register',methods=['POST','GET'])

def _register_():
    if request.method=='POST':
        course = request.form['s_course']
        name = request.form['name']
        roll_no = request.form['rollno']
        sheet=register(course,name,roll_no)
        return render_template("register.html",comment='registration success',sheet=sheet)
    return render_template("register.html")

@app.route('/list',methods=['POST','GET'])

def _list_():
    if request.method=='GET':
        list = course_list()
        return render_template("attendence.html",list=list)
    return render_template("attendence.html")

@app.route('/rlist',methods=['POST','GET'])

def _rlist_():
    if request.method=='GET':
        list = course_list()
        return render_template("register.html",list=list)
    return render_template("register.html")
    

@app.route('/a_daylist',methods=['POST','GET'])

def _aselect_():
    if request.method=='POST':
        global s_course
        s_course = request.form['s_course']
        select(s_course)
        day,day_list=a_daylist(s_course)
        return render_template("attendence.html",days=day_list)
    return render_template("attendence.html")

@app.route('/a-post',methods=['POST','GET'])

def _apost_():
    if request.method=='POST':
        global s_course
        day,day_list=a_daylist(s_course)
        a_day = request.form['a_day']
        nlist = request.form['n_list'].split()
        llist = request.form['l_list'].split()
        sheet = a_post(a_day,day,s_course,nlist,llist)
        return render_template("attendence.html",sheet=sheet,status = 'attendence posted successfully')
    return render_template("attendence.html")

@app.route('/s_daylist',methods=['POST','GET'])

def _rselect_():
    if request.method=='POST' or request.method=='GET':
        global s_course
        s_course = request.form['s_course']
        select(s_course)
        day,day_list=s_daylist(s_course)
        return render_template("attendence.html",days=day_list)
    return render_template("attendence.html")

@app.route('/s-post',methods=['POST','GET'])

def _spost_():
    if request.method=='POST':
        global s_course
        day,day_list=s_daylist(s_course)
        s_day = request.form['s_day']
        rollno = request.form['rollno']
        status = request.form['status']
        sheet = s_post(s_day,day,s_course,rollno,status)
        return render_template("attendence.html",sheet=sheet,status = 'attendence posted successfully')
    return render_template("attendence.html")

@app.route('/slist',methods=['POST','GET'])

def _slist_():
    if request.method=='GET':
        list = course_list()
        return render_template("status.html",list=list)
    return render_template("status.html")

@app.route('/a-status',methods=["GET","POST"])

def _status_():
    if request.method=='POST':
        course = request.form['s_course']
        roll = request.form['roll']
        comment,sheet=a_status(course,roll)
        return render_template("status.html",comment=comment,sheet=sheet,roll=roll)
    return render_template("status.html")





if __name__=="__main__":
    s_course = ''
    app.run(debug=True,port=34)
        
