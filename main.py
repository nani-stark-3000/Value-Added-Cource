import xlsxwriter
import openpyxl
import pandas as pd
import secrets as s
import boto3
import cv2
import os
import io

s3 = boto3.resource(
    service_name='s3',
    region_name="ap-south-1",
    aws_access_key_id= s.access_key,
    aws_secret_access_key= s.secret_access_key
)

os.environ["AWS_DEFAULT_REGION"] = s.region_name
os.environ["AWS_ACCESS_KEY_ID"] = s.access_key
os.environ["AWS_SECRET_ACCESS_KEY"] = s.secret_access_key





def create():
    print("Enter details to create sheet... ")
    course_name = input("Course name : ")
    year = input("Year : ")
    branch = input("Brach name : ")
    filename = course_name+'-'+branch+'-'+year+'.xlsx'
    print(filename,type(filename))
    wb = xlsxwriter.Workbook(filename)
    sheet1 = wb.add_worksheet("Attendance")
    sheet2 = wb.add_worksheet("Feedback")
    sheet1.write(0,0,'Name')
    sheet1.write(0,1,'Reg Number')
    sheet1.write(0,2,'Day-1')
    sheet2.write(0,1,'Course Content')
    sheet2.write(0,2,'Coverage Of Syllabus')
    sheet2.write(0,3,'Very Helpful For Your Skill Development')
    sheet2.write(0,4,'Effectiveness Of The Course')
    sheet2.write(0,5,'Interatcion & Individual Attenction')
    sheet2.write(1,0,'Excelent')
    sheet2.write(2,0,'Very Good')
    sheet2.write(3,0,'Good')
    sheet2.write(4,0,'Satisfactory')
    sheet2.write(5,0,'To Be Improved')
    for i in range(1,6):
        for j in range(1,6):
            sheet2.write(i,j,0)
    wb.close()
    update(filename)

def register(course_name):
    print("Enter details for Course Registration...")
    wb = openpyxl.load_workbook(course_name)
    ws = wb["Attendance"]
    name = input("Name : ")
    reg_number = input("Reg Number : ")
    info = (name,reg_number)
    ws.append(info)
    wb.save(course_name)
    update(course_name)

def add_day(day,course_name):
    wb = openpyxl.load_workbook(course_name)
    ws = wb['Attendance']
    new_day = 'Day-'+str(len(day)-2)
    prev_day = 'Day-'+str(len(day)-3)
    day[new_day] = chr(ord(day[prev_day])+1)
    ws[day[new_day]+'1']=new_day
    print("added",new_day)
    return day[new_day],new_day

def attendance_post(course_name):
    df = pd.read_excel(course_name)
    col = list(df.columns)
    alpha ='A'
    day = {'add':0}
    for i in col:
        day[i]=alpha
        alpha=chr(ord(alpha)+1)
    print("select day")
    day_list = [i for i in day]
    day_list.append('add')
    print(day_list[3:])
    choise = input("Enter your choise : ")
    today = day[choise]
    if today == 0 :
        today,choise=add_day(day,course_name)
    head = today+'1'
    wb = openpyxl.load_workbook(course_name)
    sheet = wb.active
    sheet[head]= choise
    nlist = input("Enter N Absenties : ").split()
    llist = input("Enter L Absenties : ").split()

    n= ['21L31A54'+i for i in nlist]
    l= ['22L35A54'+i for i in llist]
    rows = sheet.max_row
    for i in range(2,rows+1):
        roll='B'+str(i)
        atnd=today+str(i)
        if (sheet[roll].value in n) or (sheet[roll].value in l):
            sheet[atnd]= 0
        else:
            sheet[atnd]= 1 
    wb.save(course_name)
    update(course_name)

def select():
    course_list = []
    for obj in s3.Bucket('vaccourse').objects.all():
        course_list.append(obj.key)
    print(course_list)
    s_course = input("copy paste the course name : ")
    s3.Bucket('vaccourse').download_file(Key=s_course, Filename=s_course)
    #obj = s3.Bucket('vaccourse').Object(s_course).get()
    #data = obj['Body'].read()
    #df = pd.read_excel(io.BytesIO(data), index_col=0)
    #df.to_excel(s_course)
    return s_course

def update(course_name):
    s3.Bucket('vaccourse').upload_file(Filename=course_name, Key=course_name)

def special_attendance(course_name):
    df = pd.read_excel(course_name)
    col = list(df.columns)
    alpha ='A'
    day = {}
    for i in col:
        day[i]=alpha
        alpha=chr(ord(alpha)+1)
    print("select day")
    day_list = [i for i in day]
    print(day_list[2:])
    choise = input("Enter your choise : ")
    today = day[choise]
    wb = openpyxl.load_workbook(course_name)
    sheet = wb.active
    prefix = ['21L31A54','22L35A54']
    roll_no = prefix[int(input("0 for normal 1 for latral : "))]+input("Enter roll number : ")
    status = int(input("0 for absent 1 for present : "))
    rows = sheet.max_row
    for i in range(2,rows+1):
        roll='B'+str(i)
        atnd=today+str(i)
        if sheet[roll].value==roll_no:
            sheet[atnd]= status
            print("Attendance updated ")
            break
    wb.save(course_name)
    update(course_name)

def attendance_status(course_name):
    wb = openpyxl.load_workbook(course_name)
    sheet = wb.active
    prefix = ['21L31A54','22L35A54']
    roll_no = prefix[int(input("0 for normal 1 for latral : "))]+input("Enter roll number : ")
    rows = sheet.max_row
    col = sheet.max_column
    status=flag = 0
    for i in range(2,rows+1):
        roll='B'+str(i)
        if sheet[roll].value==roll_no:
            flag=1
            alpha = 'C'
            for j in range(3,col+1):
                if int(sheet[alpha+str(i)].value)==1:
                    status+=1
                alpha = chr(ord(alpha)+1)
    if flag:
        print("Your attendance is {0} for {1} days".format(status,col-2))
        percentage = (status/(col-2))*100
        print(percentage,"%")
    else:
        print("data not found with ..",roll_no)
    wb.save(course_name)
    update(course_name)

def feedback(course_name):
    print("Enter your feedback...")
    wb = openpyxl.load_workbook(course_name)
    ws = wb["Feedback"]
    openion = {'Excelent':'2','Very Good':'3','Good':'4','Satisfactory':'5','To Be Improved':'6'}
    

    list = {'1.Course Content':'B','2.Coverage of Syllabus':'C','3.Very helpful for your skill Develpoment':'D','4.Effectiveness of the course':'E','5.Interaction & Individual attention':'F'}

    
    for i in list:
        print(i)
        print([i for i in openion])
        temp = input("Enter yor openion : ")
        value = list[i]+openion[temp]
        ws[value] = ws[value].value+1

    wb.save(course_name)
    update(course_name)

600,260

def certificate():
    template = cv2.imread('student certificate.png')
    cv2.putText(template,'Konada Narendra Neeraj',(600,260),cv2.FONT_HERSHEY_COMPLEX,0.4,(0,0,255),1,cv2.LINE_AA)
    cv2.imwrite('nani.png',template)
    print("certificate generated successfully")


s_course = ''
menu = ['create','select','register','a-post','s-a-post','a-status','feedback']


while 1:
    print("select operation",menu)
    op = input(":- ")
    if op==menu[0]:
        create()
    elif op==menu[1]:
        s_course = select()
    elif op==menu[2]:
        if s_course != '':
            register(s_course)
        else:
            print("select course first...")
    elif op==menu[3]:
        if s_course != '':
            attendance_post(s_course)
        else:
            print("select course first...")
    elif op==menu[4]:
        if s_course != '':
            special_attendance(s_course)
        else:
            print("select course first...")
    elif op==menu[5]:
        if s_course != '':
            attendance_status(s_course)
        else:
            print("select course first...")
    elif op==menu[6]:
        if s_course != '':
            feedback(s_course)
        else:
            print("select course first...")
    else:
        break
